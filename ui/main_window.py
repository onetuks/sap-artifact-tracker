from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QMessageBox, QStatusBar, QFileDialog,
)
from PyQt6.QtCore import Qt

from api.client import SapClient
from api.packages_api import deploy_artifact, delete_designtime_artifact
from api.runtime_api import undeploy_artifact
from models.artifact import Artifact, ArtifactStatus
from workers.fetch_worker import FetchWorker, ActionWorker
from ui.tenant_panel import TenantPanel
from ui.tenant_dialog import TenantDialog
from ui.filter_bar import FilterBar
from ui.artifact_table import ArtifactTableWidget
from ui.confirm_dialog import ConfirmDialog, make_progress_dialog


class MainWindow(QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("SAP IS Artifact Tracker")
        self.resize(1100, 680)
        self._fetch_worker: FetchWorker | None = None
        self._action_worker: ActionWorker | None = None
        self._current_tenant = None
        self._build_ui()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setSpacing(6)
        root.setContentsMargins(12, 10, 12, 10)

        # 상단: 테넌트 패널
        self._tenant_panel = TenantPanel()
        self._tenant_panel.add_button().clicked.connect(self._on_add_tenant)
        self._tenant_panel.search_requested.connect(self._on_search)
        self._tenant_panel.tenant_selected.connect(self._on_tenant_selected)
        root.addWidget(self._tenant_panel)

        # 필터 바
        self._filter_bar = FilterBar()
        self._filter_bar.package_changed.connect(
            lambda t: self._table.proxy().set_package_filter(t)
        )
        self._filter_bar.artifact_changed.connect(
            lambda t: self._table.proxy().set_artifact_filter(t)
        )
        self._filter_bar.status_changed.connect(
            lambda s: self._table.proxy().set_status_filter(s)
        )
        root.addWidget(self._filter_bar)

        # 아티팩트 테이블
        self._table = ArtifactTableWidget()
        root.addWidget(self._table)

        # 하단 액션 버튼
        action_bar = QHBoxLayout()

        # 선택 관련 버튼
        self._select_all_btn = QPushButton("Select All")
        self._uncheck_all_btn = QPushButton("Uncheck All")
        self._select_all_btn.setFixedHeight(32)
        self._uncheck_all_btn.setFixedHeight(32)
        self._select_all_btn.clicked.connect(self._table.check_all_visible)
        self._uncheck_all_btn.clicked.connect(self._table.uncheck_all_visible)
        action_bar.addWidget(self._select_all_btn)
        action_bar.addWidget(self._uncheck_all_btn)

        action_bar.addSpacing(20)

        # 작업 관련 버튼
        self._deploy_btn = QPushButton("Deploy Selected")
        self._undeploy_btn = QPushButton("Undeploy Selected")
        self._delete_btn = QPushButton("Delete Selected")
        self._export_btn = QPushButton("Export to Excel")
        for btn in (self._deploy_btn, self._undeploy_btn, self._delete_btn, self._export_btn):
            btn.setFixedHeight(32)
            action_bar.addWidget(btn)

        self._deploy_btn.clicked.connect(lambda: self._on_action("Deploy"))
        self._undeploy_btn.clicked.connect(lambda: self._on_action("Undeploy"))
        self._delete_btn.clicked.connect(lambda: self._on_action("Delete"))
        self._export_btn.clicked.connect(self._on_export_excel)
        root.addLayout(action_bar)

        # 상태 바
        self._status_bar = QStatusBar()
        self.setStatusBar(self._status_bar)
        self._status_bar.showMessage("테넌트를 선택하고 Search를 클릭하세요.")

    # ── 테넌트 ──────────────────────────────────────────────

    def _on_add_tenant(self):
        dlg = TenantDialog(self)
        if dlg.exec() == TenantDialog.DialogCode.Accepted:
            self._tenant_panel.refresh()
            self._status_bar.showMessage("테넌트가 추가되었습니다.")

    def _on_tenant_selected(self, tenant):
        self._current_tenant = tenant

    # ── 검색 ──────────────────────────────────────────────

    def _on_search(self, tenant):
        if self._fetch_worker and self._fetch_worker.isRunning():
            return
        self._current_tenant = tenant
        self._status_bar.showMessage("아티팩트 목록을 가져오는 중...")
        self._set_buttons_enabled(False)
        self._filter_bar.clear()

        self._fetch_worker = FetchWorker(tenant, self)
        self._fetch_worker.finished.connect(self._on_fetch_done)
        self._fetch_worker.error.connect(self._on_fetch_error)
        self._fetch_worker.start()

    def _on_fetch_done(self, artifacts: list):
        self._table.load_artifacts(artifacts)
        self._status_bar.showMessage(f"총 {len(artifacts)}개의 아티팩트를 불러왔습니다.")
        self._set_buttons_enabled(True)

    def _on_fetch_error(self, msg: str):
        QMessageBox.critical(self, "조회 실패", f"아티팩트 조회 중 오류가 발생했습니다.\n\n{msg}")
        self._status_bar.showMessage("조회 실패.")
        self._set_buttons_enabled(True)

    # ── 일괄 작업 ──────────────────────────────────────────

    def _on_action(self, action: str):
        print(f"DEBUG: _on_action 호출됨 - action={action}")
        if self._current_tenant is None:
            QMessageBox.warning(self, "테넌트 없음", "테넌트를 먼저 선택하세요.")
            return

        checked = self._table.get_checked_artifacts()
        print(f"DEBUG: 선택된 아티팩트 수: {len(checked)}")
        if not checked:
            QMessageBox.information(self, "선택 없음", "작업할 아티팩트를 체크박스로 선택해주세요.")
            return

        # 작업별 사전 검증
        if action == "Deploy":
            invalid = [a for a in checked if a.status == ArtifactStatus.INACTIVE]
            if invalid:
                QMessageBox.warning(
                    self, "배포 불가",
                    "INACTIVE 상태(design-time 없음) 아티팩트는 배포할 수 없습니다.\n선택을 해제 후 다시 시도하세요.",
                )
                return
        elif action == "Undeploy":
            print(f"DEBUG: Undeploy validation 확인 중")
            invalid = [a for a in checked if a.status == ArtifactStatus.NOT_DEPLOYED]
            print(f"DEBUG: NOT_DEPLOYED 아티팩트 수: {len(invalid)}")
            print(f"DEBUG: 아티팩트 상태들: {[a.status for a in checked]}")
            if invalid:
                QMessageBox.warning(
                    self, "Undeploy 불가",
                    "NOT DEPLOYED 상태(배포된 런타임 없음) 아티팩트는 undeploy할 수 없습니다.",
                )
                return
        elif action == "Delete":
            invalid = [a for a in checked if a.status == ArtifactStatus.INACTIVE]
            if invalid:
                QMessageBox.warning(
                    self, "삭제 불가",
                    "INACTIVE 상태 아티팩트는 design-time이 없어 삭제할 수 없습니다.",
                )
                return

        dlg = ConfirmDialog(action, checked, self)
        if dlg.exec() != ConfirmDialog.DialogCode.Accepted:
            print("DEBUG: 확인 다이얼로그에서 취소됨")
            return

        print(f"DEBUG: _run_action 호출 - action={action}, 아티팩트 수={len(checked)}")
        self._run_action(action, checked)

    def _run_action(self, action: str, artifacts: list[Artifact]):
        action_fn = self._make_action_fn(action)
        progress_dlg = make_progress_dialog(action, len(artifacts), self)

        self._action_worker = ActionWorker(self._current_tenant, action_fn, artifacts, self)

        def on_progress(current, total, name):
            if progress_dlg.wasCanceled():
                self._action_worker.requestInterruption()
                return
            progress_dlg.setValue(current)
            progress_dlg.setLabelText(f"{action} 중... ({current}/{total})\n{name}")

        def on_finished(failures: list):
            progress_dlg.close()
            if failures:
                detail = "\n".join(failures)
                QMessageBox.warning(
                    self, "일부 실패",
                    f"{len(artifacts) - len(failures)}개 성공, {len(failures)}개 실패.\n\n{detail}",
                )
            else:
                QMessageBox.information(self, "완료", f"{len(artifacts)}개 {action} 완료.")
            # 목록 갱신
            self._on_search(self._current_tenant)

        self._action_worker.progress.connect(on_progress)
        self._action_worker.finished.connect(on_finished)
        self._action_worker.error.connect(
            lambda e: (progress_dlg.close(), QMessageBox.critical(self, "오류", e))
        )
        self._set_buttons_enabled(False)
        self._action_worker.start()
        progress_dlg.exec()

    def _on_export_excel(self):
        checked = self._table.get_checked_artifacts()
        if not checked:
            QMessageBox.information(self, "선택 없음", "내보낼 아티팩트를 체크박스로 선택해주세요.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Excel 파일 저장",
            "artifacts.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Artifacts"

            headers = ["Package", "Artifact", "Runtime", "Status", "Version"]
            ws.append(headers)

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for art in checked:
                ws.append([
                    art.package_name,
                    art.artifact_name,
                    art.runtime_name,
                    art.status.value,
                    art.version,
                ])

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            wb.save(file_path)
            QMessageBox.information(self, "성공", f"{len(checked)}개의 아티팩트를 Excel 파일로 내보냈습니다.\n\n{file_path}")
            self._status_bar.showMessage(f"Excel 파일로 {len(checked)}개 아티팩트를 내보냈습니다.")
        except Exception as e:
            QMessageBox.critical(self, "내보내기 실패", f"Excel 파일 생성 중 오류가 발생했습니다.\n\n{str(e)}")

    @staticmethod
    def _make_action_fn(action: str):
        """작업 종류에 따른 (SapClient, Artifact) -> None 함수 반환."""
        if action == "Deploy":
            def fn(client: SapClient, art: Artifact):
                deploy_artifact(client, art.artifact_id, art.version)
        elif action == "Undeploy":
            def fn(client: SapClient, art: Artifact):
                undeploy_artifact(client, art.artifact_id)
        elif action == "Delete":
            def fn(client: SapClient, art: Artifact):
                # Deployed 상태면 먼저 undeploy
                if art.status == ArtifactStatus.DEPLOYED:
                    undeploy_artifact(client, art.artifact_id)
                delete_designtime_artifact(client, art.artifact_id, art.version)
        else:
            raise ValueError(f"알 수 없는 작업: {action}")
        return fn

    def _set_buttons_enabled(self, enabled: bool):
        for btn in (self._select_all_btn, self._uncheck_all_btn, self._deploy_btn, self._undeploy_btn, self._delete_btn, self._export_btn):
            btn.setEnabled(enabled)
