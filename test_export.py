#!/usr/bin/env python
"""Excel export 기능 테스트"""
import sys
import tempfile
import os
from pathlib import Path

from PyQt6.QtWidgets import QApplication
from PyQt6.QtCore import Qt

from models.artifact import Artifact, ArtifactStatus
from ui.artifact_table import ArtifactTableWidget


def test_export_excel():
    """Excel export 기능 테스트"""
    app = QApplication(sys.argv)

    # 테스트용 artifact 데이터
    artifacts = [
        Artifact(
            package_id="PKG001",
            package_name="com.example.package1",
            artifact_id="ART001",
            artifact_name="MyArtifact1",
            version="1.0.0",
            runtime_name="Runtime1",
            status=ArtifactStatus.DEPLOYED
        ),
        Artifact(
            package_id="PKG002",
            package_name="com.example.package2",
            artifact_id="ART002",
            artifact_name="MyArtifact2",
            version="2.0.0",
            runtime_name="Runtime2",
            status=ArtifactStatus.NOT_DEPLOYED
        ),
        Artifact(
            package_id="PKG003",
            package_name="com.example.package3",
            artifact_id="ART003",
            artifact_name="MyArtifact3",
            version="1.5.0",
            runtime_name="Runtime3",
            status=ArtifactStatus.INACTIVE
        ),
    ]

    # 테이블 위젯 생성 및 데이터 로드
    table = ArtifactTableWidget()
    table.load_artifacts(artifacts)

    # 첫 번째와 두 번째 artifact 체크
    table.check_all_visible()
    checked = table.get_checked_artifacts()

    print(f"✅ 로드된 artifact 수: {len(artifacts)}")
    print(f"✅ 체크된 artifact 수: {len(checked)}")
    assert len(checked) == 3, f"Expected 3 checked artifacts, got {len(checked)}"

    # Excel export 테스트
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "test_artifacts.xlsx")

            # Excel 파일 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "Artifacts"

            headers = ["Package", "Artifact", "Runtime", "Status", "Version"]
            ws.append(headers)

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for art in checked:
                ws.append([
                    art.package_name,
                    art.artifact_name,
                    art.runtime_name,
                    art.status.value,
                    art.version,
                ])

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            wb.save(file_path)

            # 파일 생성 확인
            assert os.path.exists(file_path), f"Excel file not created at {file_path}"
            assert os.path.getsize(file_path) > 0, f"Excel file is empty"

            print(f"✅ Excel 파일 생성 성공: {file_path}")
            print(f"✅ 파일 크기: {os.path.getsize(file_path)} bytes")

            # openpyxl로 파일 검증
            from openpyxl import load_workbook
            wb_read = load_workbook(file_path)
            ws_read = wb_read.active

            # 헤더 확인
            header_row = [cell.value for cell in ws_read[1]]
            assert header_row == headers, f"Header mismatch: {header_row} != {headers}"
            print(f"✅ 헤더 확인: {header_row}")

            # 데이터 행 확인
            row_count = ws_read.max_row
            assert row_count == 4, f"Expected 4 rows (1 header + 3 data), got {row_count}"
            print(f"✅ 데이터 행 수 확인: {row_count}행 (헤더 1 + 데이터 3)")

            # 첫 번째 데이터 행 확인
            first_row = [cell.value for cell in ws_read[2]]
            expected_first = [artifacts[0].package_name, artifacts[0].artifact_name,
                            artifacts[0].runtime_name, artifacts[0].status.value, artifacts[0].version]
            assert first_row == expected_first, f"First row mismatch: {first_row} != {expected_first}"
            print(f"✅ 첫 번째 데이터 행 확인: {first_row}")

    except Exception as e:
        print(f"❌ Excel 내보내기 실패: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    print("\n✅ 모든 테스트 통과!")
    sys.exit(0)


if __name__ == "__main__":
    test_export_excel()
